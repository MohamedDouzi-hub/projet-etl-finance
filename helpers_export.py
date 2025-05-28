import os
from typing import Dict, Literal
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

import numpy as np
import pandas as pd
from sqlalchemy import create_engine, MetaData

IfExists = Literal["fail", "replace", "append"]

# Cette fonction prend un dictionnaire de DataFrames {nom_feuille: DataFrame}
# et les enregistre dans un fichier Excel unique, avec un onglet par DataFrame.
# Des styles personnalisés (coloration conditionnelle, mise en gras) sont appliqués
# à certaines feuilles comme 'regression' et 'mean_by_sector' pour faciliter la lecture.
# Le répertoire est créé automatiquement si nécessaire.
def dataframes_to_excel(
        dataframes: Dict[str, pd.DataFrame], excel_full_path: str
) -> None:
    """
    Exporte des DataFrames dans un fichier Excel à partir d’un dictionnaire
    où les clés correspondent aux noms des feuilles et les valeurs aux DataFrames.
    Applique des couleurs conditionnelles aux feuilles 'regression' et 'mean_by_sector'.
    """
    os.makedirs(os.path.dirname(excel_full_path), exist_ok=True)

    with pd.ExcelWriter(excel_full_path) as writer:
        for sheet, df in dataframes.items():
            if isinstance(df, pd.Series):
                df.to_frame().to_excel(writer, sheet_name=sheet, merge_cells=False, index=True)
            elif isinstance(df.columns, pd.MultiIndex):
                df.to_excel(writer, sheet_name=sheet, merge_cells=False)
            elif df.index.dtype == np.int64 and df.index.nlevels == 1:
                df.to_excel(writer, sheet_name=sheet, merge_cells=False, index=False)
            else:
                df.to_excel(writer, sheet_name=sheet, merge_cells=False, index=True)

    # Coloration conditionnelle après sauvegarde si p-value <0;05
    wb = load_workbook(excel_full_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if sheet_name == "regression":
            # Couleur verte si p-value < 0.05
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            for row in range(2, ws.max_row + 1):
                cell = ws[f"C{row}"]  # colonne C = P-value
                try:
                    if isinstance(cell.value, (float, int)) and float(cell.value) < 0.05:
                        cell.fill = green_fill
                except Exception:
                    continue
            # Mettre en gras les noms de variables
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
                for cell in row:
                    cell.font = Font(bold=True)

        if sheet_name == "mean_by_sector":
            # Colorier en bleu le plus grand rendement journalier et en orange la plus faible volatilité
            blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
            orange_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

            # Trouver les colonnes
            headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
            col_return = get_column_letter(headers.get("Return"))
            col_volatility = get_column_letter(headers.get("Volatility"))

            # Récupération des valeurs
            returns = [ws[f"{col_return}{i}"].value for i in range(2, ws.max_row + 1)]
            volatilities = [ws[f"{col_volatility}{i}"].value for i in range(2, ws.max_row + 1)]

            try:
                max_return = max(filter(lambda x: x is not None, returns))
                min_volatility = min(filter(lambda x: x is not None, volatilities))
            except ValueError:
                continue

            for i in range(2, ws.max_row + 1):
                if ws[f"{col_return}{i}"].value == max_return:
                    ws[f"{col_return}{i}"].fill = blue_fill
                if ws[f"{col_volatility}{i}"].value == min_volatility:
                    ws[f"{col_volatility}{i}"].fill = orange_fill

    # Sauvegarde finale
    wb.save(excel_full_path)


def dataframes_to_db(
        dataframes: Dict[str, pd.DataFrame],
        db_path: str,
        drop_all_tables: bool = False,
        append_data: bool = False,
) -> None:
    """
    Export DataFrames to a SQLite database
    :param dataframes: DataFrames as Dict(table_name, DataFrame)
    :param db_path: full path of SQLite database
    :param drop_all_tables: if True, drop all existing tables
    :param append_data: if True, add to existing tables instead of replacing
    """
    path, _ = os.path.split(db_path)
    os.makedirs(path, exist_ok=True)

    engine = create_engine(f"sqlite:///{db_path}", echo=True)
    con = engine.connect()
    meta = MetaData()

    if drop_all_tables:
        meta.reflect(bind=engine)
        meta.drop_all(bind=engine)

    if_exists: IfExists = "append" if append_data else "replace"

    for sheet_name, df in dataframes.items():
        df.to_sql(name=sheet_name, con=con, if_exists=if_exists, index=False)


def to_db_format(name: str) -> str:
    """
    Formate un nom de colonne pour le rendre compatible avec SQL
    (suppression des espaces, mise en minuscules).
    """
    return name.strip().lower().replace(" ", "_")
